from django_filters import rest_framework as rest_framework_filters
from django.db.models import F
from libs import EWSReader
from collections import OrderedDict
from apps.channel import models as channel_models
from apps.feeder import models as feeder_models
from apps.consumer import models as consumer_models
from apps.report import serializers
from apps.report import filters

from datetime import datetime
from datetime import timedelta
from dateutil.relativedelta import relativedelta
from copy import deepcopy

from openpyxl import Workbook
from openpyxl import styles
from openpyxl import utils
#from tempfile import TemporaryFile
from wsgiref.util import FileWrapper
from uuid import uuid1 as uuid
from os import unlink


class DataReportMixin:
	Model = channel_models.Channel
	serializer_class = serializers.Report
	queryset = Model.objects.all()
	filterset_class = filters.DataReportFilter

	request_keys = (
		'req',
		'interval',
		'tariff',
		't1',
		't2',
	)
	filter_backends = (
		rest_framework_filters.DjangoFilterBackend,
	)

	def get_crq_kwargs(self):
		if not hasattr(self, '_crq_kwargs'):
			self._crq_kwargs = OrderedDict()
			for key in self.request_keys:
				if key not in self.request.GET:
					return
				self._crq_kwargs[key] = self.request.GET.get(key)

		return self._crq_kwargs

	def get_queryset_kwargs(self):
		if not hasattr(self, '_qs_kwargs'):
			self._qs_kwargs = {
				key: self.request.GET[key]
				for key in self.request.GET
				if key not in self.request_keys and self.request.GET[key]
			}
		return self._qs_kwargs

	def filter_queryset(self, queryset):
		kwargs = self.get_crq_kwargs()

		if kwargs:
			ews = EWSReader(
				dataset=queryset,
				const_params=kwargs,
			)

			return ews.full_response()
		else:
			return queryset

	def get_queryset(self):
		if self.get_crq_kwargs():
			kwargs = self.get_queryset_kwargs()

			return self.queryset\
				.annotate(
					consumer_id=F('feeder__consumer_id'),
					channel_id=F('id'),
					const_datapath=F('datapath__const_datapath__path'),
					variable_datapath=F('datapath__variable_datapath'),
				)\
				.values(
					'consumer_id',
					'feeder_id',
					'channel_id',
					'const_datapath',
					'variable_datapath',
					'datapath',
					'channel_type',
				)\
				.filter(**kwargs)\
				.order_by(
					'datapath__const_datapath__path',
					'datapath__variable_datapath',
				)
		else:
			return self.queryset.none()


class GroupedReportMixin(DataReportMixin):
	serializer_class = serializers.GroupedReport

	@staticmethod
	def parse_crq_req_date(req_date):
		return datetime.strptime(
			req_date,
			'%Y%m%d%H%M%S',
		)

	intervals = {
		'short': timedelta(minutes=3),
		'main': timedelta(minutes=30),
		'day': timedelta(days=1),
		'month': relativedelta(months=1),
		'year': relativedelta(years=1),
	}

	def build_time_serie(self):
		crq_kwargs = self.get_crq_kwargs()
		interval = self.intervals[crq_kwargs['interval']]
		t1 = self.parse_crq_req_date(crq_kwargs['t1'])
		t2 = self.parse_crq_req_date(crq_kwargs['t2'])
		if crq_kwargs['interval'] in {'main'}:
			t1 += interval
			t2 += interval
		t2 += interval
		while t1 < t2:
			yield t1
			t1 += interval

	def build_row_sketeton(self, queryset):
		consumer_row = []

		index = {
			'consumer': {},
			'feeder': {},
			'channel': {},
		}

		for row in queryset:
			consumer_id = row['consumer_id']

			if consumer_id in index['consumer']:
				consumer_index = index['consumer'][consumer_id]
				consumer = consumer_row[consumer_index]
				feeder_row = consumer['feeder']
			else:
				feeder_row = []
				consumer = {
					'id': consumer_id,
					'feeder': feeder_row,
				}
				index['consumer'][consumer_id] = len(consumer_row)
				consumer_row.append(consumer)

			feeder_id = row['feeder_id']

			if feeder_id in index['feeder']:
				feeder_index = index['feeder'][feeder_id]
				feeder = feeder_row[feeder_index]
				channel_row = feeder['channel']
			else:
				channel_row = []
				feeder = {
					'id': feeder_id,
					'channel': channel_row,
				}
				index['feeder'][feeder_id] = len(feeder_row)
				feeder_row.append(feeder)

			channel_id = row['channel_id']
			channel_type = row['channel_type']

			if channel_id in index['channel']:
				channel_index = index['channel'][channel_id]
				value = channel_row[channel_index]
				value_record = value['channel']
			else:
				value_record = {
					'value': None,
					'valid': False,
				}
				value = {
					'id': channel_id,
					'type': channel_type,
					'value': value_record,
				}
				index['channel'][channel_id] = len(channel_row)
				channel_row.append(value)

		return {
			'consumer': consumer_row,
			'index': index,
		}

	def group_table_iter(self, row_skeleton):
		consumer = row_skeleton['consumer']
		index = row_skeleton['index']
		index['date'] = {}

		for date_index, date in enumerate(self.build_time_serie()):
			index['date'][date] = date_index
			yield {
				'date': date,
				'consumer': deepcopy(consumer),
			}


	def mk_grouped_table(self, queryset):
		row_skeleton = self.build_row_sketeton(queryset)
		consumer = row_skeleton['consumer']
		index = row_skeleton['index']
		return {
			'table': tuple(self.group_table_iter(row_skeleton)),
			'index': index,
			'consumer': consumer,
		}

	def fill_grouped_table(self, queryset):
		if not self.get_crq_kwargs():
			return {
				'table': (),
				'index': {},
				'consumer': {},
			}

		grouped_table = self.mk_grouped_table(queryset)
		table = grouped_table['table']
		index = grouped_table['index']

		for row in super().filter_queryset(queryset):
			date = row['date']
			consumer_id = row['consumer_id']
			feeder_id = row['feeder_id']
			channel_id = row['channel_id']

			date_index = index['date'][date]
			date_row = table[date_index]

			consumer_index = index['consumer'][consumer_id]
			consumer_row = date_row['consumer'][consumer_index]

			feeder_index = index['feeder'][feeder_id]
			feeder_row = consumer_row['feeder'][feeder_index]

			channel_index = index['channel'][channel_id]
			channel_row = feeder_row['channel'][channel_index]
			value = channel_row['value']
			for key in ('value', 'valid'):
				value[key] = row[key]

		return grouped_table

	def filter_queryset(self, queryset):
		queryset = tuple(queryset)
		grouped_table = self.fill_grouped_table(queryset)
		return grouped_table['table']


class FileGroupedReportMixin(GroupedReportMixin):
	font_error = styles.Font(
		color=styles.colors.RED,
		bold=True,
	)
	serializer_class = serializers.FileSerializer
	feeder_queryset = feeder_models.Feeder.objects.all()
	consumer_queryset = consumer_models.Consumer.objects.all()

	def write_headers(self, grouped_table, wsheet):
		date_column = 1
		date_letter = utils.get_column_letter(date_column)
		wsheet.column_dimensions[date_letter].width = 20

		consumer = grouped_table['consumer']
		consumer_row_number = 1
		feeder_row_number = 2
		channel_row_number = 3
		column = 1

		value_cell = wsheet.cell(
			row=channel_row_number,
			column=column,
		)

		value_cell.value = 'Date'

		column += 1
		for consumer_index, consumer_row in enumerate(consumer):
			consumer_start_col = column
			feeder = consumer_row['feeder']
			for feeder_index, feeder_row in enumerate(feeder):
				feeder_start_col = column
				channel = feeder_row['channel']
				for channel_index, channel_row in enumerate(channel):
					value_cell = wsheet.cell(
						row=channel_row_number,
						column=column,
					)
					channel_type = channel_row['type']
					channel_choise = channel_models.Channel.choices[channel_type]
					channel_name = channel_choise[1]
					value_cell.value = channel_name
					column += 1

				feeder_end_col = column - 1
				cell_range = '{startcol}{row}:{endcol}{row}'.format(
					startcol=utils.get_column_letter(feeder_start_col),
					endcol=utils.get_column_letter(feeder_end_col),
					row=feeder_row_number,
				)
				wsheet.merge_cells(cell_range)
				value_cell = wsheet.cell(
					row=feeder_row_number,
					column=feeder_start_col,
				)
				feeder_id = feeder_row['id']
				feeder_instance = self.feeder_queryset.get(id=feeder_id)
				value_cell.value = feeder_instance.name


			consumer_end_col = column - 1
			cell_range = '{startcol}{row}:{endcol}{row}'.format(
				startcol=utils.get_column_letter(consumer_start_col),
				endcol=utils.get_column_letter(consumer_end_col),
				row=consumer_row_number,
			)
			wsheet.merge_cells(cell_range)
			value_cell = wsheet.cell(
				row=consumer_row_number,
				column=consumer_start_col,
			)
			consumer_id = consumer_row['id']
			consumer_instance = self.consumer_queryset.get(id=consumer_id)
			value_cell.value = consumer_instance.name

	def fill_excel(self, grouped_table, wsheet, row=1):
		for time_row in grouped_table['table']:
			column = 1
			date = time_row['date']
			date_cell = wsheet.cell(row=row, column=column)
			date_cell.value = date
			consumer = time_row['consumer']
			for consumer_index, consumer_row in enumerate(consumer):
				feeder = consumer_row['feeder']
				for feeder_index, feeder_row in enumerate(feeder):
					channel = feeder_row['channel']
					for channel_index, channel_row in enumerate(channel):
						column += 1
						value = channel_row['value']
						value_cell = wsheet.cell(row=row, column=column)
						if value['valid']:
							value_cell.value = value['value']
						else:
							value_cell.value = '<null>'
							value_cell.font = self.font_error
			row += 1


	def mk_excel(self, queryset):
		grouped_table = self.fill_grouped_table(queryset)
		wb = Workbook()
		ws = wb.active
		self.write_headers(grouped_table, ws)
		self.fill_excel(grouped_table, ws, row=4)
		tmp = uuid()
		tmp = "%s.xlsx" % tmp
		wb.save(tmp)
		#filedata = b64encode(open(tmp, 'rb').read())
		filedata = open(tmp, 'rb')
		filedata = FileWrapper(filedata)
		unlink(tmp)
		return filedata

	def filter_queryset(self, queryset):
		queryset = tuple(queryset)
		return self.mk_excel(queryset)
